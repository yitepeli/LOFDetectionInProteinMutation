# -*- coding: utf-8 -*-
"""
Created on Wed Dec  5 21:28:30 2018

@author: yitepeli
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import csv

#wb = Workbook()

#dest_filename = 'empty_book.xlsx'

# klmnopq

KEYS= [['K','A'],['L','B'],['M','C'],['N','D'],['O','E'],['P','F'],['Q','G']]

def main():
    
    mainWB = load_workbook('PPData.xlsx')
    #print(mainWB.sheetnames)
    mainS = mainWB.active
    #if mainS['K'][6].value != "NA":
        #print(mainS['K'][6].value)


#with open('PPData.csv', 'w', newline='') as csvfile:

    with open('PPData.csv', 'w', newline='') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',)
        for row in mainS:
            spamwriter.writerow([col.value for col in row])
        #spamwriter.writerow(['Spam', 'Lovely Spam', 'Wonderful Spam'])

'''
    for rowNo in range(len(mainS[KEYS[0][0]])-3):
        mainRowIndex = rowNo+3
        if mainS[KEYS[0][0]][mainRowIndex].value != "NA":
            for key in KEYS:
                pos = key[1] + str(ppRowNo)
                #print(pos)
                #print(mainRowIndex)
                ppS1[pos] = mainS[key[0]][mainRowIndex].value
            ppRowNo += 1
            
    
    
    
    ppWB.save(filename = ppWB_filename)'''
    
if __name__== "__main__":
    main()