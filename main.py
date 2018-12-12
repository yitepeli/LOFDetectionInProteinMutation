# -*- coding: utf-8 -*-
"""
Created on Wed Dec  5 13:57:54 2018

@author: yitepeli
"""

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

#wb = Workbook()

#dest_filename = 'empty_book.xlsx'

# klmnopq

KEYS= [['K','A'],['L','B'],['M','C'],['N','D'],['O','E'],['P','F'],['Q','G']]

def main():
    
    mainWB = load_workbook('maindata.xlsx')
    #print(mainWB.sheetnames)
    mainS = mainWB.active
    #if mainS['K'][6].value != "NA":
        #print(mainS['K'][6].value)
    
    
    '''ppWB = Workbook()
    ppWB_filename = 'PPData.xlsx'
    ppS1 = ppWB.active

    ppS1['A1'] = "aa_pos"
    ppS1['B1'] = "aa_ref"
    ppS1['C1'] = "aa_alt"
    ppS1['D1'] = "protein_var"
    ppS1['E1'] = "consequence"
    ppS1['F1'] = "score"
    ppS1['G1'] = "class"
    
    ppRowNo = 2
    for rowNo in range(len(mainS[KEYS[0][0]])-3):
        mainRowIndex = rowNo+3
        if mainS[KEYS[0][0]][mainRowIndex].value != "NA":
            for key in KEYS:
                pos = key[1] + str(ppRowNo)
                #print(pos)
                #print(mainRowIndex)
                ppS1[pos] = mainS[key[0]][mainRowIndex].value
            ppRowNo += 1
            
    
    
    
    ppWB.save(filename = ppWB_filename)'''
    
if __name__== "__main__":
    main()